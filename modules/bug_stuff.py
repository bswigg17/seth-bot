from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import modules.styles as styles


def do_bug_work(sprint, bugs, week, new_sprint):
    if new_sprint:
        clean_ws(sprint)
    id_hash = make_hash(bugs)
    merge(sprint, id_hash, week)


def clean_ws(sprint):
    i = 1
    for row in enumerate(sprint):
        sprint.delete_rows(i)
    data = ('Title', 'ID', 'Week 1', 'Week 2', 'Week 3')
    sprint.append(data)


def format_value(cell, val):
    cell.font = styles.FONT 
    cell.border = styles.BORDER
    cell.alignment = styles.ALIGNMENT
    if val in {'Ready', 'In Progress', ''}:
        cell.fill = styles.FILL_WAITING
        cell.value = 'Awaiting Resolution'
    elif val == 'Test':
        cell.fill = styles.FILL_TESTING
        cell.value = 'Testing'
    elif val in {'Done', 'Accepted'}:
        cell.fill = styles.FILL_RESOLVED
        cell.value = 'RESOLVED'
    elif val == 'NEW':
        cell.value = val


# def clean_ws(ws):
#     """Iterate over row D and delete rows with resolved"""
#     row_number = 1 
#     for status in ws['D']:
#         if status.value == 'RESOLVED':
#             completed_ids.add(ws[f'B{row_number}'].value)
#             ws.delete_rows(row_number)
#             continue
#         row_number += 1

#     """Move Over Row D to Row C"""
#     ws.move_range(f"D2:D{row_number}", cols=-1)


def make_hash(ws_export):
    """Make Set of ID's for quick lookup"""
    id_hash = {}
    i = 1
    for id_ in ws_export['B']:
        #Check for ID Row and None Row
        if id_.value == 'ID':
            i += 1
            continue
        # elif id_.value in completed_ids:
        #     i += 1
        #     continue
        #Check for Duplicate Rows
        elif id_.value in id_hash:
            ws.delete_rows(i)
            continue
        id_hash[id_.value] = {'Title': ws_export[f'A{i}'].value, 'Status': ws_export[f'D{i}'].value}
        i += 1
    return id_hash


def format_new_row(ws, week, id_, data, row): 
    ws[f'A{row}'] = data['Title']
    ws[f'B{row}'] = id_ 
    i = ord('C')
    while i < ord(week):
        format_value(ws[f'{chr(i)}{row}'], 'NEW')
        i += 1
    format_value(ws[f'{week}{row}'], data['Status'])


def merge(ws, id_hash, week):
    i = 1
    for id_ in ws['B']:
        if id_.value == 'ID':
            i += 1
            continue
        elif id_.value in id_hash:
            format_value(ws[f'{week}{i}'], id_hash[id_.value]['Status'])
            del id_hash[id_.value]
            i += 1
        else:
            ws.delete_rows(i)

    # ws.delete_rows(i+1)
    if len(id_hash) > 0:
        for id_ in id_hash:
            format_new_row(ws, week, id_, id_hash[id_], i)
            i += 1

# if __name__ == "__main__":
#     wb, bug_ws = load_workbook("modules/Sprint Status Elera Pay MC (5).xlsx"), load_workbook("modules/Export (6).xlsx").active
#     ws = wb['Defect Status']
#     do_bug_work(ws, bug_ws, chr(66 + 1), 1)
#     wb.save('New.xlsx')