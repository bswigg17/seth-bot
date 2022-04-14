from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.chart import BarChart, Series, Reference, PieChart
import modules.styles as styles


def default_styling(cell_):
    cell_.font = styles.FONT
    cell_.border = styles.BORDER
    cell_.alignment = styles.ALIGNMENT


def format_status(status):
    if status == 'Accepted' or status == 'Done':
        return 'Completed'
    elif status ==  'In Progress':
        return 'Likely'
    elif status == 'Test':
        return 'Testing'
    elif status == 'Ready' or status == '':
        return 'TBD'
    elif status == 'Moved':
        return 'Moved'
    elif status == 'New':
        return 'NEW'


def format_data(status, cell_):
    default_styling(cell_)
    #Match Logic
    if status == 'Carry':
        cell_.fill = styles.FILL_CARRIED_OVER
        cell_.value = 'Carried Over'
    elif status == 'Completed':
        cell_.fill = styles.FILL_COMPLETED
        cell_.value = status
    elif status ==  'Likely':
        cell_.fill = styles.FILL_LIKELY
        cell_.value = status
    elif status == 'Testing':
        cell_.fill = styles.FILL_TESTING
        cell_.value = status
    elif status == 'TBD':
        cell_.fill = styles.FILL_TBD
        cell_.value = status
    elif status == 'Moved':
        cell_.fill = styles.FILL_MOVED
        cell_.value = "MOVED"
    elif status == 'New':
        cell_.value = 'NEW'


def format_row(ws, id_, title, status, row, new=False):
    default_styling(ws[f'A{row}'])
    ws[f'A{row}'].value = title
    default_styling(ws[f'B{row}'])
    ws[f'B{row}'].value = id_
    if not new:
        ws[f'B{row}'].fill = styles.FILL_CARRIED_OVER
    format_data(status, ws[f'C{row}']) 


def notMicahTessa(val):
    if val in ['Micah Clark', 'Tessa Swier', 'Micah Clark;Tessa Swier']:
        return False
    return True
    

def add_storie_status_to_dict(row, storie_status_hash, data_hash):
    if notMicahTessa(row[2].value):
        """{ID: {Title: str, Status: str}}"""
        title = row[0].value
        id_ = row[1].value
        status = format_status(row[-2].value)
        owner = row[2].value
        points = int(row[-1].value) if row[-1].value != "" else 0
        """Make entry into hash"""
        storie_status_hash[id_] = {'Title': title, 'Status': status, 'Owner': owner, 'Points': points}
           
        """{Status: {Likely: int, Completed: int}}"""
        if status in data_hash:
            data_hash[status]['Total'] += 1
            data_hash[status]['Points'] += points


def create_data_hash():
    status_types = ['TBD', 'Likely', 'Testing', 'Completed', 'Moved']
    data_dict = dict()
    for status in status_types:
        data_dict[status] = {'Points': 0, 'Total': 0}
    return data_dict

def extract_storie_status_from_ws_backlog(ws_export):
    storie_status_hash = dict()
    data_hash = create_data_hash()
    i = 1
    for row in ws_export:
        if i <= 1:
            i += 1
            continue 
        else:
            add_storie_status_to_dict(row, storie_status_hash, data_hash)
            i += 1
    return storie_status_hash, data_hash


def format_new_row(ws, data, id_, row_num, column):
    ws[f'A{row_num}'] = data['Title']
    ws[f'B{row_num}'] = id_

    i = ord('C')
    while i < ord(column):
        format_data('New', ws[f'{chr(i)}{row_num}'] )
        i += 1
    format_data(data['Status'], ws[f'{column}{row_num}'])


def row_is_header(id_value):
    return True if id_value == 'ID' else False


def row_is_duplicate(id_value, done_ids):
    return True if id_value in done_ids else False


def handle_row_edge_case(ws, ws_row_num):
    return ws.delete_rows(ws_row_num)


def story_is_now_completed(status):
    print(status)
    return True if status == 'Completed' else False

def story_completed_last_week(ws, column, row):
    return True if ws[f'{chr(ord(column) - 1)}{row}'].value == 'Completed' else False


def story_was_moved(id_, id_hash):
    return True if id_ not in id_hash else False


def update_row_for_new_sprint(ws, id_, id_hash, done_ids, column, ws_row_num, new_sprint_ws, new_ws_row_num):
    try:
        if id_ in id_hash:
            format_data('Carry', ws[f'{column}{ws_row_num}'])
            format_row(new_sprint_ws, id_, id_hash[id_]['Title'], id_hash[id_]['Status'], new_ws_row_num)
    except Exception as e:
        print(e)


def update_rows_for_new_sprint(ws, id_hash, column, new_sprint_ws):
    ws_row_num, new_ws_row_num = 1, 1
    done_ids = set()
    for id_ in ws['B']:
        if row_is_header(id_.value):
            ws_row_num += 1
        elif row_is_duplicate(id_.value, done_ids):
            handle_row_edge_case(ws, ws_row_num)
            ws_row_num += 1
        elif id_.value in id_hash:
            format_data('Carry', ws[f'{column}{ws_row_num}'])
            format_row(new_sprint_ws, id_.value, id_hash[id_.value]['Title'], id_hash[id_.value]['Status'], new_ws_row_num)
            done_ids.add(id_.value)
            del id_hash[id_.value]
            ws_row_num += 1
            new_ws_row_num += 1 
        elif story_completed_last_week(ws, column, ws_row_num):
            format_data('Completed', ws[f'{column}{ws_row_num}'])
            ws_row_num += 1
        elif story_was_moved(id_.value, id_hash):
            format_data('Moved', ws[f'{column}{ws_row_num}'])
            ws_row_num += 1
    for id_ in id_hash:
        format_row(new_sprint_ws, id_, id_hash[id_]['Title'], id_hash[id_]['Status'], new_ws_row_num, new=True)
        new_ws_row_num += 1 


def update_rows_for_sprint(ws, id_hash, column):
    ws_row_num = 1
    done_ids = set()
    for id_ in ws['B']:
        if row_is_header(id_.value):
            ws_row_num += 1
        elif row_is_duplicate(id_.value, done_ids):
            handle_row_edge_case(ws, ws_row_num)
            ws_row_num += 1
        elif id_.value not in id_hash:
            format_data('Moved', ws[f'{column}{ws_row_num}'])
            ws_row_num += 1
        else:
            format_data(id_hash[id_.value]['Status'], ws[f'{column}{ws_row_num}'])
            done_ids.add(id_.value)
            del id_hash[id_.value]
            ws_row_num += 1
    for id_ in id_hash:
        format_new_row(ws, id_hash[id_], id_, ws_row_num, column)
        ws_row_num += 1 


def update_sprint_status(ws, id_hash, week, new_sprint_ws):
    if new_sprint_ws is not None:
        update_rows_for_new_sprint(ws, id_hash, week, new_sprint_ws)
    else:
        update_rows_for_sprint(ws, id_hash, week)


def calculate_statistics(hash_, data_hash):
    # owner_stats = {entry['Owner']: int(entry['Points']) for entry in hash_.values() if entry}
    # total_points = sum(owner_stats.values())
    # points_complete = sum([int(entry['Points']) for entry in hash_.values() if entry['Status'] == 'Accepted' or entry['Status'] == 'Done'])
    # progress = str(round(points_complete / total_points * 100)) + "%"
    # owner_stats_sorted = {k: v for (k,v) in sorted(owner_stats.items(), key=lambda x: x[-1], reverse=True)}
    data_wb = Workbook()
    data_ws = data_wb.active

    rows = [(k, v['Points'], v['Total']) for k,v in data_hash.items()]
    rows.insert(0, ('Status Type', 'Story Points Total', 'Story Count Total'))

    for row in rows:
        data_ws.append(row)
    
    chart = BarChart()
    chart.type = 'col'
    chart.style = 10
    chart.title = 'Points Per Status Type'
    chart.x_axis.title = 'Status Type'
    chart.y_axis.title = 'Total Points'
    # labels = Reference(data_ws, min_col=1, min_row=2, max_row=6)
    data = Reference(data_ws, min_col=2, min_row=1, max_row=6)
    categories = Reference(data_ws, min_col=1, min_row=2, max_row=6)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    data_ws.add_chart(chart, 'A15')

    chart2 = BarChart()
    labels_2 = Reference(data_ws, min_col=1, min_row=2, max_row=6)
    data_2 = Reference(data_ws, min_col=3, min_row=1, max_row=6)
    chart2.add_data(data_2, titles_from_data=True)
    chart2.set_categories(labels_2)
    data_ws.add_chart(chart2, 'I15')
    data_wb.save('Data.xlsx')



    
    # data_ws['A1'] = 'Owner(s)'
    # data_ws['B1'] = 'Points Completed'
    # data_ws['C1'] = "Story Status"
    # data_ws['D1'] = "Count"
    # data_ws['E1'] = "% Stories Completed"


    # for row, (owner, points) in enumerate(owner_stats_sorted.items(), 2):
    #     data_ws[f'A{row}'] = owner
    #     data_ws[f'B{row}'] = points

    
    
    # data_ws['C2'] = progress



def do_sprint_work(sprint, backlog, week, new_sprint_ws=None):
    """Make Workbook and Worksheet Objects."""
    """Extract Data From ws_export Into Dictinary"""
    try:
        storie_status_hash, data_hash = extract_storie_status_from_ws_backlog(backlog)
        calculate_statistics(storie_status_hash, data_hash)
        update_sprint_status(sprint, storie_status_hash, week, new_sprint_ws)
    except Exception as E:
        print(E)


# if __name__ == "__main__":
#     wb_sprint, wb_backlog = load_workbook('modules/ELERA Pay Sprint Status 2022.xlsx'), load_workbook('modules/Export-13.xlsx')
#     do_sprint_work(wb_sprint.active, wb_backlog.active, chr(66 + 2))
#     wb_sprint.save('New.xlsx')