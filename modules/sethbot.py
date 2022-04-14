from modules.sprint_stuff import do_sprint_work
from modules.bug_stuff import do_bug_work
from openpyxl import Workbook, load_workbook

def week_to_column(week_num):
    return chr(66 + week_num)

def load_workbooks(sprint, bugs, backlog):
    try:
        wb_sprint = load_workbook(sprint)
        wb_bugs = load_workbook(bugs)
        wb_export = load_workbook(backlog)
        return wb_sprint, wb_bugs, wb_export
    except Exception as E:
        print(E)

def seth_bot(sprint, bugs, backlog, week, new_sprint=0):
    try:
        wb_sprint, wb_bugs, wb_backlog = load_workbooks(sprint, bugs, backlog)
        sprint_week = week_to_column(week)
        if new_sprint:
            do_sprint_work(wb_sprint['Q1-22 Sprint 1'], wb_backlog.active, sprint_week, wb_sprint.create_sheet('New Sprint'))
        else:
            do_sprint_work(wb_sprint['Q1-22 Sprint 1'], wb_backlog.active, sprint_week)
        do_bug_work(wb_sprint['Defect Status'], wb_bugs.active, sprint_week, new_sprint)
        wb_sprint.save('Update.xlsx')
        return True
    except Exception as E:
        print(E)
        return False

if __name__ == "__main__":
    seth_bot()