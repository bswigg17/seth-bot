from __future__ import annotations
from openpyxl import Workbook, load_workbook
from typing import List
import csv, os
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText
import smtplib
from datetime import date
from typing import Optional, List
import pickle

# class EmailSender:

#     __sender: str
#     __sender_login: str
#     __sender_password: str
#     __recipients: List[str]
#     __subject: str
#     __body: str
#     __file_path: str
#     __msg: MIMEMultipart
#     __last_sent = None

#     def __init__(self, to_, from_="", subject="", body="", file_path= ""):
#         self.__sender = from_
#         self.__recipients = to_
#         self.__subject = subject
#         self.__body = body
#         self.__file_path = file_path

#     @property
#     def file_path(self):
#         return self.__file_path

#     @file_path.setter
#     def file_path(self, file):
#         self.__file_path = file

#     @staticmethod
#     def sendEmail(self):
#         self.__prepareEmail()
#         try: 
#         # Create SMTP object
#             smtp_obj = smtplib.SMTP('smtp.gmail.com', 587)
#             smtp_obj.starttls()
#             # Login to the server
#             smtp_obj.login('bswigg17', passwordHash())

#             # Convert the message to a string and send it
#             smtp_obj.sendmail(msg['From'], msg['To'], msg.as_string())
#             smtp_obj.quit()
#         except Exception as e:
#             print(e)
#         finally:
#             if self.__file != "":
#                 os.remove(self.__file)
#             self.__msg = MIMEMultipart()

#     @staticmethod
#     def __prepareEmail(self):
#         """Prepare Email MIMEMultipart Object For Sending"""
#         self.__configureMIME()

#     @staticmethod
#     def __configureMIME(self):
#         """Adds Information To MIMEMultipart"""
#         self.__msg['Subject'] = self.__subject
#         self.__msg['From'] = self.__from
#         self.__toList()
#         self.__attachBody()
#         self.__attachFile()

#     @staticmethod
#     def __toList(self):
#         if len(self.__recipients) >= 1:
#             self.__msg['To'] = ", ".join(self.__recipients)

#     @staticmethod
#     def __attachBody(self):
#         self.__msg.attach(MIMEText(self.__body, 'plain'))

#     @staticmethod
#     def __attachFile(self): 
#         if __file_path != "":
#             with open(self.__file_path,'rb') as file:
#                 self.__msg.attach(MIMEApplication(file.read(), Name=f"SprintStatus-{date.today()}.csv"))

#     def __repr__(self):
#         return f"EmailSender(Recipients={self.__recipients}, Sender={self.__sender}, Subject={self.__subject}, Body={self.__body}, File_Path={self.__file_path}, Last_Sent={self.__last_sent})"

#     def __str__(self):
#         return f"EmailSender(Recipients={self.__recipients}, Sender={self.__sender}, Subject={self.__subject}, Body={self.__body}, File_Path={self.__file_path}, Last_Sent={self.__last_sent})"

class ExcelConfig:
    # Instance Varibles:
        # controller: Dict[str, list[set]]
            # title_row_letter
            # id_row_letter
            # owner_row_letter
            # status_row_letter
            # points_row_letter

    def __init__(self):
        self.set_controller()

    def set_controller(self):
        self.controller = dict({
            'Title': False,
            'ID': False,
            'Owner': False,
            'Status': False,
            'Points': False,

        })


    pass


class SprintConfig(ExcelConfig):
    
    def __init__(self):
        super().__init__()


class BugConfig(ExcelConfig):

    def __init__(self):
        super().__init__()


class ExcelFile:

    # Instance Variables:
        # file_path: str
        # file_object: Workbook

    def __init__(self, file_path):
        # self.set_and_create_workbook(file_path)
        self.set_and_create_workbook(file_path)
        
    @classmethod
    def set_and_create_workbook(self, file_path):
        """Sets file path, creates workbook object from path, and grabs active worksheet of workbook"""
        self.set_file_path(file_path)
        self.create_workbook()

    @classmethod
    def set_file_path(self, file):
        self.file_path = file

    @classmethod
    def create_workbook(self):
        self.set_file_object(load_workbook(self.get_file_path()))

    # @classmethod
    # def get_active_worksheet(self):
    #     self.worksheet = self.file_object.active

    # @property
    # def worksheet():
    #     return self.worksheet

    @classmethod
    def get_file_object(self):
        return self.file_object

    @classmethod
    def set_file_object(self, file_object):
        self.file_object = file_object

    @classmethod
    def get_file_path(self):
        return self.file_path

    @classmethod
    def __repr__(self):
        return f"SethBot({self.file_path})"

    @classmethod
    def __str__(self):
         return f"SethBot: {self.file_path}"


class ExportFile(ExcelFile):

    # Instance Variables:
        # super().file_path
        # super().file_object
        # status_ws
        # id_hash

        # id_row_number
        # onwer_row_number
        # title_row_number
        # status_row_number
        # points_row_number

    def __init__(self, file_path):
        """Creates ExcelFile via super() and initializes id_hash instance variable"""
        super().__init__(file_path)
        self.set_id_hash()

    @classmethod
    def set_id_hash(self):
        self.id_hash = dict()

    @classmethod
    def get_id_hash(self):
        return self.id_hash

    @classmethod
    def fill_out_id_hash(self):
        i = 1
        for row in ws_export:
            if i <= 1:
                i += 1
                continue
            else:
                self.extract_row_data(row)

    @classmethod
    def extract_row_data(self, row):
        # if story_owner_is_tracked(row[self.owner_row_number]):
        add_row_to_id_hash(self, row)

        
    @classmethod
    def add_row_to_id_hash(self, row):
        # self.id_hash[row[self.id_row_number]] = {
        #                                             'Title': row[self.title_row_number],
        #                                             'Status': row[self.status_row_number],
        #                                             'Points': row[self.points_row_number],
        #                                         }
        self.id_hash[row[1].value] = {'Title': row[0].value,
                                'Status': row[-2].value}
        


class Status(ExcelFile):

    # sprint_ws
    # bug_ws

    def __init__(self, file_path):
        super().__init__(file_path)
        self.set_worksheets()
        print(self.bug_ws)

    @classmethod
    def set_worksheets(self):
        self.set_sprint_ws()
        self.set_bug_ws()
    
    @classmethod
    def set_sprint_ws(self):
        try:
            self.sprint_ws = self.get_file_object()['Q4-21 Sprint 3']
        except Exception as E:
            print(E)

    @classmethod
    def get_sprint_ws(self):
        return self.sprint_ws

    @classmethod
    def set_bug_ws(self):
        try:
            self.bug_ws = self.get_file_object()['Defect Status']
        except Exception as E:
            print(E)

    @classmethod
    def get_bug_ws(self):
        return self.bug_ws

    # @property
    # def sprint_ws(self):
    #     return self.sprint_ws

    # @sprint_ws.setter
    # def sprint_ws(self, ws_name):
    #     for sheet in self.file_object.worksheets:
    #         print(sheet)
    #     self.sprint_ws = self.file_object[ws_name]

    # @property
    # def bug_ws(self):
    #     return self.bug_ws

    # @bug_ws.setter
    # def bug_ws(self, ws_name):
    #     self.bug_ws = self.file_object[ws_name]
    


class Bugs(ExportFile):

    # super().file_path
    # super().file_object
    # status_ws
    # id_hash

    # done_ids = {}
    # sprint_ws 
    # bugs_export_ws

    def __init__(self, file_path, ws):
        super().__init__(file_path)
        self.set_bug_status_ws(ws)
    
    @classmethod
    def set_bug_status_ws(self, ws):
        self.bug_status_ws = ws
        self.bug_ws = self.get_file_object().active

    @classmethod
    def generate_bug_status(self):
        self.load_done_ids()
        self.append_to_status()
        self.save_done_ids()

    @classmethod
    def get_bug_ws(self):
        return self.bug_status_ws

    @classmethod
    def append_to_status(self):
        i = 1
        for id_ in (self.get_bug_ws())['B']:
            print(id_.value)
            if id_.value == 'ID':
                i += 1
                continue
            if id_.value in self.id_hash and id_value not in self.done_ids:
                self.format_value(self.status_ws[f'D{i}'], self.id_hash[id_.value]['Status'])
                self.status_wsws[f'E{i}'] = self.id_hash[id_.value]['Points']
                self.done_ids.add(id_.value)
                del self.id_hash[id_.value]
                i += 1
            else:
                (self.get_bug_ws()).delete_rows(i)

        if len(self.id_hash) > 0:
            for id_ in self.id_hash:
                if id_ in self.done_ids:
                    continue
                else:
                    data = id_hash[id_]
                    self.get_bug_ws()[f'A{i}'] = data['Title']
                    self.get_bug_ws()[f'B{i}'] = id_ 
                    format_value(self.get_bug_ws()[f'C{i}'], 'NEW')
                    format_value(self.get_bug_ws()[f'D{i}'], data['Status'])
                    self.get_bug_ws()[f'E{i}'] = data['Points']
                    i += 1
    
    @classmethod
    def format_value(self, cell, val):
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
        if val in {'Ready', 'In Progress', ''}:
            cell.fill = fill_waiting
            cell.value = 'Awaiting Resolution'
        elif val == 'Test':
            cell.fill = fill_testing
            cell.value = 'Testing'
        elif val in {'Done', 'Accepted'}:
            cell.fill = fill_resolved
            cell.value = 'RESOLVED'
        elif val == 'NEW':
            cell.value = val

    @classmethod
    def load_done_ids(self):
        try: 
            with open('done_ids', 'rb') as f:
                self.done_ids = pickle.load(f)
        except:
            self.done_ids = set()

    @classmethod
    def save_done_ids(self):
        with open('done_ids', 'wb') as f:
            pickle.dump(self.done_ids, f)


class Backlog(ExportFile):

    def __init__(self, file_path, sprint_ws):
        super().__init__(file_path)


    @classmethod
    def generate_backlog_status(self):
        pass

    @classmethod
    def extract_row_data(self, row):
        if story_owner_is_tracked(row[self.owner_row_number]):
            add_row_to_id_hash(self, row)

    @classmethod
    def story_owner_is_tracked(self, owner):
        if val in ['Micah Clark', 'Tessa Swier', 'Micah Clark;Tessa Swier']:
            return False
        return True

    @classmethod
    def append_rows_to_status(self, sprint_ws):
        i = 1
        for id_ in ws['B']:
            if id_.value == 'ID':
                i += 1
                continue
            if id_.value in self.id_hash:
                format_data(id_hash[id_.value]['Status'], ws[f'{column}{i}'])
                done_ids.add(id_.value)
                del id_hash[id_.value]
            elif id_.value not in id_hash and id_.value not in done_ids:
                format_data('Moved', ws[f'{column}{i}'])
            elif id_.value in done_ids:
                ws.delete_rows(i)
                continue
            i += 1
        for id_ in id_hash:
            format_new_row(ws, id_hash[id_], id_, i, column)
            i+= 1 

    @classmethod
    def append_row_to_status(self, row_data):
        pass


class SethBot: 
    """Controls All Excel Logic and Emails The Resulting Excel File Each Week"""
    # __status: Status
    # __bugs: Bugs
    # __backlog: Backlog
    # __email: Email
    # __week: int
    # __config: ExcelConfig

    def __init__(self, status: str, bugs: str, backlog: str, week: int):
        self.set_fields(status, bugs, backlog, week)
        self.do_work()

    @classmethod
    def set_fields(self, status: str, bugs: str, backlog: str, week: int):
        self.set_status(status)
        self.set_bugs(bugs)
        self.set_backlog(backlog)
        self.set_week(week)

    @classmethod 
    def do_work(self):
        self.__bugs.generate_bug_status()
        #self.__backlog.generate_backlog_status()

    @classmethod
    def new_week(self):
        pass
    
    @classmethod
    def set_status(self, status_file_path):
        try:
            self.__check_file(status_file_path)
            self.__status = Status(status_file_path)
            print(self.__status)
        except Exception as e:
            print(e)

    @classmethod
    def set_bugs(self, bug_file_path):
        try:
            self.__check_file(bug_file_path)
            self.__bugs = Bugs(bug_file_path, self.__status.get_bug_ws())
        except Exception as e:
            print(e)

    @classmethod
    def set_backlog(self, backlog_file_path):
        try:
            self.__check_file(backlog_file_path)
            self.__backlog = Backlog(backlog_file_path, self.__status.sprint_ws)
        except Exception as e:
            print(e)

    @classmethod
    def set_week(self, week):
        if isinstance(week, int):
            self.__week = week

    @classmethod
    def __check_file(self, file):
        file_ending = file.split(".")[-1]
        if file_ending == "xlsx":
            return
        raise Exception("Not Xlsx")

    @classmethod
    def update_week(self):
        self.__week = (self.__week + 1) % 4

    @classmethod
    def __repr__(self):
        return f"SethBot({self.__status}, {self.__bugs}, {self.__backlog}, {self.__week}, {self.__email})"

    @classmethod
    def __str__(self):
         return f"SethBot: {self.__status}, {self.__bugs}, {self.__backlog}, {self.__week}, {self.__email}"