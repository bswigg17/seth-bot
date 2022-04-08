from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

FONT = Font(name='Arial',
            size=11,
            bold=True, 
            )

side = Side(border_style='thin')

BORDER = Border(top=side,bottom=side, left=side, right=side)

ALIGNMENT = Alignment(horizontal='center')

FILL_RESOLVED = PatternFill(fill_type='solid',
                            start_color='0000B050',
                            end_color='0000B050')
                            

FILL_WAITING = PatternFill(fill_type='solid',
                            start_color='A9D08E',
                            end_color='A9D08E')

FILL_TESTING = PatternFill(fill_type='solid',
                            start_color='0000B0F0',
                            end_color='0000B0F0')

FILL_CARRIED_OVER  = PatternFill(fill_type='solid',
                    start_color='FFC000',
                    end_color='FFC000')

FILL_TBD = PatternFill(fill_type='solid',
                    start_color='BDD7EE',
                    end_color='BDD7EE')

FILL_LIKELY = PatternFill(fill_type='solid',
                    start_color='92D050',
                    end_color='92D050')

FILL_MOVED = PatternFill(fill_type='solid',
                    start_color='FF0000',
                    end_color='FF0000')

FILL_TESTING = PatternFill(fill_type='solid',
                    start_color='00B0F0',
                    end_color='00B0F0')

FILL_COMPLETED = PatternFill(fill_type='solid',
                    start_color='00B050',
                    end_color='00B050')
